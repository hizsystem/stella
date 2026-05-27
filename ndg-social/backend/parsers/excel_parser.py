"""
ExcelParser — NDG 소셜 광고 데이터 Excel → JSON 변환

지원 컬럼 (자동 매핑 + 유연 매핑):
  발행일 / 게재 위치 / 좋아요 / 댓글 / 저장
  총 인터랙션 / 노출 / 프로필 방문 / 팔로우 / 광고 예산
"""

import re
from datetime import datetime
from io import BytesIO
from typing import Optional

import openpyxl
import pandas as pd


COLUMN_ALIASES = {
    "upload_date":        ["업로드일", "날짜", "date", "upload_date", "게시일", "발행일", "발행 일"],
    "title":              ["제목", "콘텐츠명", "콘텐츠 제목", "게시물 제목", "내용", "title",
                           "post_title", "게시물", "소재명", "콘텐츠", "게재내용", "소재"],
    "content_type":       ["콘텐츠 구분", "콘텐츠구분", "구분", "type", "content_type",
                           "게재 위치", "게재위치"],
    "content_format":     ["유형", "포맷", "형식", "format"],
    "ad_placement":       ["광고 위치", "광고위치", "게재위치2", "placement", "ad_placement"],
    "likes":              ["좋아요", "likes", "like"],
    "comments":           ["댓글", "comments", "comment"],
    "saves":              ["저장", "saves", "save", "북마크"],
    "total_interactions": ["총 인터랙션", "총인터랙션", "인터랙션", "interactions", "total_interactions"],
    "impressions":        ["조회수", "노출수", "impressions", "views", "노출"],
    "reach":              ["도달", "reach"],
    "profile_visits":     ["프로필 방문", "프로필방문", "profile_visits"],
    "shares":             ["공유", "shares", "share"],
    "reposts":            ["리포스트", "reposts", "repost"],
    "new_followers":      ["팔로우", "팔로워 증가", "follows", "new_followers"],
    "ad_spend":           ["광고비", "ad_spend", "spend", "비용", "광고 예산", "광고예산"],
}

CONTENT_TYPE_MAP = {
    "피드": "feed", "feed": "feed", "카드뉴스": "feed",
    "스토리": "story", "story": "story",
    "릴스": "reel", "reel": "reel", "reels": "reel",
    "피드 & 릴스": "feed", "피드&릴스": "feed",
    "카루셀": "feed", "carousel": "feed",
    "영상": "reel",
    "단일 이미지": "feed", "단일이미지": "feed",
}

EVENT_KEYWORDS = ["이벤트", "event", "증정", "기념", "giveaway", "퀴즈", "경품"]

# TOTAL 행 판별 키워드 (normalize_col 결과 기준 — 소문자, 공백 제거)
SKIP_ROW_KEYWORDS = {"total", "합계", "소계", "총계", "누계", "subtotal", "grandtotal"}


def _normalize_col(name: str) -> str:
    return str(name).strip().lower().replace(" ", "")


def _map_columns(df: pd.DataFrame) -> dict[str, str]:
    normalized = {_normalize_col(c): c for c in df.columns}
    mapping = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            key = _normalize_col(alias)
            if key in normalized:
                mapping[field] = normalized[key]
                break
    return mapping


def _safe_int(val) -> int:
    try:
        if pd.isna(val):
            return 0
        return int(float(str(val).replace(",", "").replace("원", "").strip()))
    except (ValueError, TypeError):
        return 0


def _safe_float(val) -> float:
    try:
        if pd.isna(val):
            return 0.0
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _parse_content_type(raw: str) -> tuple[str, str]:
    raw = str(raw).strip()
    raw_lower = raw.lower()
    ctype = "feed"
    # 정확 일치 우선
    if raw_lower in {k.lower(): k for k in CONTENT_TYPE_MAP}:
        for k, v in CONTENT_TYPE_MAP.items():
            if k.lower() == raw_lower:
                ctype = v
                break
    else:
        # 부분 일치 — 긴 키부터 매칭 (예: "피드 & 릴스"가 "피드"보다 먼저)
        for k, v in sorted(CONTENT_TYPE_MAP.items(), key=lambda x: len(x[0]), reverse=True):
            if k.lower() in raw_lower:
                ctype = v
                break
    subtype = "event" if any(kw in raw for kw in EVENT_KEYWORDS) else "organic"
    return ctype, subtype


def _parse_date(date_raw, report_month: str) -> Optional[str]:
    """날짜 파싱 — '2월 6일', datetime, '2026-02-06' 등 처리"""
    if pd.isna(date_raw):
        return None
    date_str = str(date_raw).strip()
    if not date_str:
        return None

    # TOTAL / 헤더 행 필터 (정규화 후 비교)
    if _normalize_col(date_str) in SKIP_ROW_KEYWORDS:
        return None
    # 날짜처럼 보이지 않는 긴 문자열(소재명 등)은 스킵
    if len(date_str) > 20 and not re.search(r'\d', date_str):
        return None

    # '2월 6일' 형식 처리
    m = re.match(r"(\d{1,2})월\s*(\d{1,2})일", date_str)
    if m:
        year = report_month.split("-")[0]
        return f"{year}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"

    # datetime 객체
    if isinstance(date_raw, datetime):
        return date_raw.strftime("%Y-%m-%d")

    # pandas Timestamp
    try:
        return pd.to_datetime(date_raw).strftime("%Y-%m-%d")
    except Exception:
        pass

    return None


def _find_header_row(file_bytes: bytes, sheet_name) -> int:
    """발행일/날짜 등이 포함된 실제 헤더 행 인덱스 반환"""
    raw = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None)
    date_keys = {_normalize_col(a) for a in COLUMN_ALIASES["upload_date"]}

    for i in range(min(15, len(raw))):
        row_vals = {_normalize_col(str(v)) for v in raw.iloc[i] if pd.notna(v)}
        if row_vals & date_keys:
            return i
    return 0


def _build_column_names(file_bytes: bytes, sheet_name, header_row: int) -> list[str]:
    """
    헤더가 2행인 경우(병합 + 서브헤더) 하위 행의 컬럼명을 우선 사용.
    e.g. row N: [NO, 발행일, ..., 인게이지먼트(오가닉), ...]
         row N+1: [..., ..., ..., 노출, 도달, 좋아요, ...]
    """
    raw = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None)
    row1 = raw.iloc[header_row]

    all_field_aliases = {_normalize_col(a) for aliases in COLUMN_ALIASES.values() for a in aliases}

    # 다음 행도 헤더인지 확인
    if header_row + 1 < len(raw):
        row2 = raw.iloc[header_row + 1]
        row2_vals = {_normalize_col(str(v)) for v in row2 if pd.notna(v) and str(v).strip()}
        if row2_vals & all_field_aliases:
            # 두 행 합치기: row2 값이 있으면 우선, 없으면 row1 값
            combined = []
            for j in range(max(len(row1), len(row2))):
                v1 = str(row1.iloc[j]).strip() if j < len(row1) and pd.notna(row1.iloc[j]) else ""
                v2 = str(row2.iloc[j]).strip() if j < len(row2) and pd.notna(row2.iloc[j]) else ""
                # row2가 실제 컬럼명이면 우선
                if v2 and _normalize_col(v2) in all_field_aliases:
                    combined.append(v2)
                elif v1 and _normalize_col(v1) in all_field_aliases:
                    combined.append(v1)
                elif v2:
                    combined.append(v2)
                elif v1:
                    combined.append(v1)
                else:
                    combined.append(f"_col_{j}")
            return combined

    return [str(v).strip() if pd.notna(v) else f"_col_{i}" for i, v in enumerate(row1)]


def _calc_summary(posts: list[dict]) -> dict:
    if not posts:
        return {}
    return {
        "total_likes":         sum(p.get("likes", 0) for p in posts),
        "total_comments":      sum(p.get("comments", 0) for p in posts),
        "total_saves":         sum(p.get("saves", 0) for p in posts),
        "total_shares":        sum(p.get("shares", 0) for p in posts),
        "total_reposts":       sum(p.get("reposts", 0) for p in posts),
        "total_interactions":  sum(p.get("total_interactions", 0) for p in posts),
        "total_impressions":   sum(p.get("impressions", 0) for p in posts),
        "total_reach":         sum(p.get("reach", 0) for p in posts),
        "total_profile_visits":sum(p.get("profile_visits", 0) for p in posts),
        "total_new_followers": sum(p.get("new_followers", 0) for p in posts),
        "total_ad_spend":      sum(p.get("ad_spend", 0) for p in posts),
    }


def _calc_ad_breakdown(posts: list[dict]) -> dict:
    breakdown: dict[str, dict] = {
        "feed":  {"spend": 0, "impressions": 0, "interactions": 0},
        "story": {"spend": 0, "impressions": 0, "views": 0},
        "reel":  {"spend": 0, "impressions": 0, "views": 0},
    }
    for p in posts:
        ctype = p.get("content_type", "feed")
        if ctype not in breakdown:
            ctype = "feed"
        breakdown[ctype]["spend"]       += p.get("ad_spend", 0)
        breakdown[ctype]["impressions"] += p.get("impressions", 0)
        if ctype == "feed":
            breakdown[ctype]["interactions"] += p.get("total_interactions", 0)
        else:
            breakdown[ctype]["views"] += p.get("impressions", 0)

    for key in ("story", "reel"):
        b = breakdown[key]
        if b["views"] > 0:
            b["cpv"] = round(b["spend"] / b["views"])
    feed = breakdown["feed"]
    if feed["interactions"] > 0:
        feed["cpc"] = round(feed["spend"] / feed["interactions"])

    return breakdown


_AD_TABLE_HEADER_KEYS = {"광고기간", "집행비용", "집행 비용", "광고목표", "광고 목표"}

_AD_COL_ALIASES: dict[str, list[str]] = {
    "title":      ["소재", "콘텐츠명", "게재내용", "콘텐츠"],
    "ad_type":    ["유형", "포맷"],
    "ad_period":  ["광고기간"],
    "objective":  ["광고목표", "광고 목표"],
    "target":     ["타겟"],
    "spend":      ["집행비용", "집행 비용", "광고비", "광고 예산"],
    "impressions":["노출"],
    "reach":      ["도달"],
    "cpm":        ["CPM", "cpm"],
    "action":     ["Action", "action", "참여", "게시물참여"],
    "atr":        ["ATR", "atr"],
    "cpa":        ["CPA", "cpa"],
    "clicks":     ["Click", "click", "링크클릭"],
    "cpc":        ["CPC", "cpc"],
    "ctr":        ["CTR", "ctr"],
}


def _parse_ad_table(raw: "pd.DataFrame", report_month: str = "") -> list[dict]:
    """
    Sheet 내 광고 집행 테이블 섹션 파싱 (story_extra TOTAL 이후 위치).
    헤더 조건: 동일 행에 _AD_TABLE_HEADER_KEYS 중 2개 이상 포함.
    """
    ad_header_row = None
    for i in range(len(raw)):
        row_vals_norm = {_normalize_col(str(v)) for v in raw.iloc[i] if pd.notna(v) and str(v).strip()}
        matches = sum(1 for kw in _AD_TABLE_HEADER_KEYS if _normalize_col(kw) in row_vals_norm)
        if matches >= 2:
            ad_header_row = i
            break

    if ad_header_row is None:
        return []

    header_row_data = raw.iloc[ad_header_row]
    col_map: dict[str, int] = {}
    for col_idx, cell_val in enumerate(header_row_data):
        if pd.isna(cell_val):
            continue
        norm = _normalize_col(str(cell_val))
        if not norm:
            continue
        for field, aliases in _AD_COL_ALIASES.items():
            if any(_normalize_col(a) == norm or norm.startswith(_normalize_col(a)) for a in aliases):
                if field not in col_map:
                    col_map[field] = col_idx
                break

    title_idx = col_map.get("title")
    if title_idx is None:
        return []

    ads: list[dict] = []
    for i in range(ad_header_row + 1, len(raw)):
        row = raw.iloc[i]
        row_vals_norm = {_normalize_col(str(v)) for v in row if pd.notna(v) and str(v).strip()}
        if row_vals_norm & SKIP_ROW_KEYWORDS:
            break

        title_val = row.iloc[title_idx] if title_idx < len(row) else None
        title = str(title_val).strip() if pd.notna(title_val) and str(title_val).strip() not in ("", "nan") else ""
        if not title:
            continue

        entry: dict = {"title": title}
        for field, idx in col_map.items():
            if field == "title":
                continue
            val = row.iloc[idx] if idx < len(row) else None
            is_null = val is None or (isinstance(val, float) and pd.isna(val))
            if is_null:
                entry[field] = 0 if field in ("spend", "impressions", "reach") else ""
            elif field in ("spend", "impressions", "reach"):
                entry[field] = _safe_int(val)
            else:
                s = str(val).strip()
                entry[field] = "" if s == "nan" else s
        ads.append(entry)

    return ads


def parse_excel(
    file_bytes: bytes,
    report_month: str,
    client: str = "HIZ-NDG",
    followers_start: int = 0,
    followers_end: int = 0,
    previous_summary: Optional[dict] = None,
    sheet_name: int | str = 0,
) -> dict:
    """Excel 파일 → raw_excel JSON 구조 반환"""

    # 1. 헤더 행 자동 감지
    header_row = _find_header_row(file_bytes, sheet_name)

    # 2. 컬럼명 구성 (2행 헤더 처리 포함)
    col_names = _build_column_names(file_bytes, sheet_name, header_row)

    # 3. 데이터 읽기 — 헤더 행 + 서브헤더 행 이후부터
    raw = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None)
    # 데이터 시작 행 결정
    data_start = header_row + 1
    # 혹시 서브헤더 행이 있으면 한 줄 더 건너뜀
    all_field_aliases = {_normalize_col(a) for aliases in COLUMN_ALIASES.values() for a in aliases}
    if data_start < len(raw):
        next_row_vals = {_normalize_col(str(v)) for v in raw.iloc[data_start] if pd.notna(v)}
        if next_row_vals & all_field_aliases:
            data_start += 1

    df = raw.iloc[data_start:].reset_index(drop=True)
    df.columns = col_names[:len(df.columns)]
    df.dropna(how="all", inplace=True)

    # 4. 컬럼 매핑
    col_map = _map_columns(df)

    # upload_date가 없으면 에러
    if "upload_date" not in col_map:
        available = list(df.columns)
        raise ValueError(
            f"날짜 컬럼을 찾을 수 없습니다. "
            f"'발행일', '업로드일', '날짜' 중 하나가 있어야 합니다. "
            f"현재 컬럼: {available}"
        )

    # 5. 행 파싱 (두 번째 테이블 감지 포함)
    # 전략: TOTAL 행(합계) 이후 등장하는 첫 유효 데이터 행부터 story_extra
    _date_col = col_map.get("upload_date", "")
    _date_aliases_norm = {_normalize_col(a) for a in COLUMN_ALIASES["upload_date"]}

    posts = []
    _table_section = "main"
    _past_total = False          # TOTAL 행을 지났는지 여부
    _last_date = f"{report_month}-01"  # carry-forward용 마지막 유효 날짜
    for idx, row in df.iterrows():
        # TOTAL 행 감지: 행 내 임의 셀에 합계/total 키워드
        _row_vals_norm = {_normalize_col(str(v)) for v in row.values if pd.notna(v) and str(v).strip()}
        if _row_vals_norm & SKIP_ROW_KEYWORDS:
            _past_total = True
            continue  # TOTAL 행 자체는 스킵

        # TOTAL 이후 & 아직 story_extra 아니면 → story_extra로 전환
        if _past_total and _table_section == "main":
            # 날짜 헤더 행이나 섹션 제목 행은 건너뜀
            _dv = _normalize_col(str(row.get(_date_col, "")).strip())
            if _dv in _date_aliases_norm or (not _dv and not any(pd.notna(v) and str(v).strip() for v in row.values)):
                continue
            _table_section = "story_extra"

        date_raw = row.get(col_map.get("upload_date", ""), "")
        upload_date = _parse_date(date_raw, report_month)
        if not upload_date:
            # NO 컬럼이 숫자이면 날짜 없이도 유효한 데이터 행으로 처리
            # 셀 병합으로 날짜가 비어있는 경우 이전 행 날짜(carry-forward) 사용
            no_col_val = ""
            for cname in ("NO", "no", "No", "번호"):
                if cname in row.index:
                    no_col_val = str(row[cname]).strip()
                    break
            try:
                float(no_col_val)  # 숫자면 유효한 행
                upload_date = _last_date  # 이전 행 날짜 승계
            except (ValueError, TypeError):
                continue
        else:
            _last_date = upload_date  # 유효한 날짜 갱신

        title_raw = row.get(col_map.get("title", ""), "")
        title = str(title_raw).strip() if title_raw and str(title_raw).strip() not in ("nan", "") else ""

        content_raw = row.get(col_map.get("content_type", ""), "")
        ctype, subtype = _parse_content_type(str(content_raw))

        # 유형(포맷) 컬럼: "단일", "카루셀" 등 — content_format으로 별도 보존
        fmt_col = col_map.get("content_format", "")
        content_format_raw = str(row.get(fmt_col, "")).strip() if fmt_col else ""
        content_format = content_format_raw if content_format_raw not in ("nan", "") else ""

        # 광고 위치(ad_placement) 컬럼에 "스토리"가 있으면 스토리로 오버라이드
        placement_col = col_map.get("ad_placement", "")
        ad_placement_raw = str(row.get(placement_col, "")).strip() if placement_col else ""
        if ad_placement_raw and ("스토리" in ad_placement_raw or "story" in ad_placement_raw.lower()):
            ctype = "story"

        likes        = _safe_int(row.get(col_map.get("likes", ""), 0))
        comments     = _safe_int(row.get(col_map.get("comments", ""), 0))
        saves        = _safe_int(row.get(col_map.get("saves", ""), 0))
        shares       = _safe_int(row.get(col_map.get("shares", ""), 0))
        reposts      = _safe_int(row.get(col_map.get("reposts", ""), 0))
        total_int    = _safe_int(row.get(col_map.get("total_interactions", ""), 0))
        if total_int == 0:
            total_int = likes + comments + saves + shares + reposts
        impressions    = _safe_int(row.get(col_map.get("impressions", ""), 0))
        reach          = _safe_int(row.get(col_map.get("reach", ""), 0))
        profile_visits = _safe_int(row.get(col_map.get("profile_visits", ""), 0))
        new_followers  = _safe_int(row.get(col_map.get("new_followers", ""), 0))
        ad_spend       = _safe_int(row.get(col_map.get("ad_spend", ""), 0))
        # 원본 게재 위치 텍스트 보존 (ad_placement 우선, 없으면 content_type 컬럼 값)
        if ad_placement_raw and ad_placement_raw not in ("nan", ""):
            placement = ad_placement_raw
        else:
            placement_raw = row.get(col_map.get("content_type", ""), "")
            placement = str(placement_raw).strip() if placement_raw and str(placement_raw).strip() not in ("nan", "") else ""

        # 모든 메트릭이 0이고 제목/위치도 비어있으면 빈 행 — 스킵
        if not title and not placement and total_int == 0 and impressions == 0:
            continue

        posts.append({
            "upload_date":        upload_date,
            "title":              title,
            "content_type":       ctype,
            "content_format":     content_format,
            "content_placement":  placement,
            "content_subtype":    subtype,
            "likes":              likes,
            "comments":           comments,
            "saves":              saves,
            "shares":             shares,
            "reposts":            reposts,
            "total_interactions": total_int,
            "impressions":        impressions,
            "reach":              reach,
            "profile_visits":     profile_visits,
            "new_followers":      new_followers,
            "ad_spend":           ad_spend,
            "is_boosted":         ad_spend > 0,
            "table_section":      _table_section,
        })

    # 합계 행 후처리: 인터랙션/도달/노출 중 하나라도 나머지 전체 합보다 크면 합계 행 판정·제거
    if len(posts) > 1:
        def _is_total_row_outlier(p: dict, all_posts: list) -> bool:
            for metric in ("total_interactions", "reach", "impressions"):
                total = sum(x.get(metric, 0) for x in all_posts)
                val = p.get(metric, 0)
                if val > 0 and val > total - val:
                    return True
            return False
        original_posts = list(posts)
        posts = [p for p in posts if not _is_total_row_outlier(p, original_posts)]

    summary = _calc_summary(posts)
    summary["followers_start"] = followers_start
    summary["followers_end"] = followers_end if followers_end > 0 else (
        followers_start + summary.get("total_new_followers", 0)
    )

    ad_breakdown = _calc_ad_breakdown(posts)
    ad_table     = _parse_ad_table(raw, report_month)

    return {
        "meta": {
            "report_month": report_month,
            "client": client,
            "parsed_at": datetime.utcnow().isoformat(),
            "row_count": len(posts),
        },
        "current_month": {
            "period": {
                "start": f"{report_month}-01",
                "end":   f"{report_month}-28",
            },
            "summary": summary,
            "posts": posts,
        },
        "previous_month": previous_summary,
        "ad_breakdown": ad_breakdown,
        "ad_table": ad_table,
    }


def parse_kpi_sheet(file_bytes: bytes, sheet: int | str = 1) -> list[dict]:
    """
    KPI 추이 시트 파싱 — 월별 KPI 추이 반환.
    sheet: 시트 인덱스(int) 또는 시트 이름(str), 기본값 1 (두 번째 시트)

    구조:
      Row 3  — 월 레이블 (col7=6월, col8=7월, ..., col15=2월 ...)
      Row 4  — 팔로워 누적 (col6='실적(누적)')
      Row 6  — 게시물 수 (col6='게시물 수')
      Row 7  — 인터랙션 합계 (col6='인터랙션')
      Row 8  — 도달 (col6='도달')
      Row 9  — 광고 예산 (col6='광고 예산')
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        if isinstance(sheet, int):
            if sheet >= len(wb.worksheets):
                return []
            ws = wb.worksheets[sheet]
        else:
            if sheet not in wb.sheetnames:
                return []
            ws = wb[sheet]

        # 목표값 (row4 col5 = 70000)
        target_raw = ws.cell(4, 5).value
        try:
            target = int(float(str(target_raw).replace(",", "").strip())) if target_raw else 0
        except (ValueError, TypeError):
            target = 0

        # 월 레이블 매핑: col_index → label ("7월" 등)
        month_cols: dict[int, str] = {}
        for cell in ws[3]:
            if cell.value and isinstance(cell.value, str) and "월" in cell.value:
                month_cols[cell.column] = cell.value

        if not month_cols:
            return {"target": target, "trend": []}

        # 지표 행 매핑
        data_rows: dict[str, dict[int, float]] = {
            "followers": {}, "achievement": {}, "post_count": {},
            "interactions": {}, "reach": {}, "ad_spend": {},
        }

        label_field_map = {
            "실적":   "followers",     # row4 col6 = "실적(누적)"
            "달성률": "achievement",   # row5 col6 = "달성률(목표대비)"
            "게시물": "post_count",    # row6 col6 = "게시물 수"
            "인터랙": "interactions",  # row7 col6 = "인터랙션"
            "도달":   "reach",         # row8 col6 = "도달"
            "광고":   "ad_spend",      # row9 col6 = "광고 예산"
        }

        for row_idx in range(4, 12):
            label_val = str(ws.cell(row_idx, 6).value or "").strip()
            # row4에서 col2에 "인스타그램" 있으면 팔로워 행 특별 처리
            if row_idx == 4 and not label_val:
                label_val = str(ws.cell(row_idx, 2).value or "").strip()

            field = None
            for kw, f in label_field_map.items():
                if kw in label_val:
                    field = f
                    break
            if not field:
                continue

            for col_idx in month_cols:
                raw = ws.cell(row_idx, col_idx).value
                if raw is None or raw == "-":
                    continue
                try:
                    if field == "achievement":
                        val = float(str(raw).replace(",", "").strip())
                    else:
                        val = int(float(str(raw).replace(",", "").strip()))
                    data_rows[field][col_idx] = val
                except (ValueError, TypeError):
                    pass

        trend = []
        for col_idx in sorted(month_cols.keys()):
            label       = month_cols[col_idx]
            followers   = data_rows["followers"].get(col_idx, 0)
            achievement = data_rows["achievement"].get(col_idx, 0)
            post_cnt    = data_rows["post_count"].get(col_idx, 0)
            inter       = data_rows["interactions"].get(col_idx, 0)
            avg_inter   = round(inter / post_cnt) if post_cnt else 0
            reach       = data_rows["reach"].get(col_idx, 0)
            ad_spend    = data_rows["ad_spend"].get(col_idx, 0)

            # 팔로워도 달성률도 없으면 미래/미반영 월 — 스킵
            if followers == 0 and achievement == 0:
                continue

            trend.append({
                "label":            label,
                "followers":        followers,
                "achievement_pct":  round(achievement * 100, 1) if achievement else 0,
                "interactions":     inter,
                "avg_interactions": avg_inter,
                "reach":            reach,
                "ad_spend":         ad_spend,
            })

        return {"target": target, "trend": trend}
    except Exception:
        return {"target": 0, "trend": []}


def get_unmapped_columns(file_bytes: bytes, sheet_name: int | str = 0) -> dict:
    header_row = _find_header_row(file_bytes, sheet_name)
    col_names  = _build_column_names(file_bytes, sheet_name, header_row)
    raw = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None)
    df  = raw.iloc[header_row + 1:].reset_index(drop=True)
    df.columns = col_names[:len(df.columns)]
    col_map  = _map_columns(df)
    mapped   = list(col_map.values())
    unmapped = [c for c in df.columns if c not in mapped and not c.startswith("_col_")]
    return {"mapped": col_map, "unmapped": unmapped, "all_columns": col_names}
