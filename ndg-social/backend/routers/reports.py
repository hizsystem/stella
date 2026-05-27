import io
import json
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Body
from fastapi.responses import StreamingResponse
import openpyxl
from sqlalchemy.orm import Session

from backend.database import Report, ReportData, ReportVersion, SlideEdit, get_db
from backend.schemas import ReportCreate, ReportOut, SlideUpdateRequest, VersionCreate
from backend.routers.auth import get_current_user
from backend.database import User
from backend.parsers.excel_parser import parse_excel, get_unmapped_columns, parse_kpi_sheet

router = APIRouter(prefix="/api/reports", tags=["reports"])


def _build_excel_bytes(data: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "운영 데이터"
    ws.append(["날짜", "콘텐츠 유형", "서브유형", "좋아요", "댓글", "저장",
               "인터랙션", "노출수", "도달수", "프로필 방문", "신규 팔로워", "광고비", "광고여부"])
    for post in data.get("current_month", {}).get("posts", []):
        ws.append([
            post.get("upload_date", ""),
            post.get("content_type", ""),
            post.get("content_subtype", ""),
            post.get("likes", 0),
            post.get("comments", 0),
            post.get("saves", 0),
            post.get("total_interactions", 0),
            post.get("impressions", 0),
            post.get("reach", 0),
            post.get("profile_visits", 0),
            post.get("new_followers", 0),
            post.get("ad_spend", 0),
            "Y" if post.get("is_boosted") else "N",
        ])
    ws2 = wb.create_sheet("요약")
    summary = data.get("current_month", {}).get("summary", {})
    ws2.append(["항목", "값"])
    for k, v in summary.items():
        ws2.append([k, v])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.post("", response_model=ReportOut)
def create_report(body: ReportCreate, db: Session = Depends(get_db),
                  current_user: User = Depends(get_current_user)):
    report = Report(
        title=body.title,
        client_name=body.client_name,
        report_month=body.report_month,
        status="draft",
        created_by=current_user.id,
    )
    db.add(report)
    db.commit()
    db.refresh(report)
    return report


@router.get("", response_model=list[ReportOut])
def list_reports(db: Session = Depends(get_db),
                 current_user: User = Depends(get_current_user)):
    return db.query(Report).order_by(Report.id.desc()).all()


@router.get("/kpi-summary")
def get_kpi_summary(db: Session = Depends(get_db),
                    current_user: User = Depends(get_current_user)):
    """
    모든 raw_excel 레코드의 kpi_trend를 합산 — 월별로 가장 최근 값 사용.
    새 파일을 업로드해도 전체 누적이 유지됨.
    """
    def _label_to_month(label: str) -> str:
        try:
            m = int(label.replace("월", "").strip())
            year = 2025 if m >= 6 else 2026
            return f"{year}-{m:02d}"
        except Exception:
            return label

    # 모든 raw_excel을 오래된 순으로 순회 → 최신 값이 덮어씌움
    all_raws = (
        db.query(ReportData)
        .filter(ReportData.data_type == "raw_excel")
        .order_by(ReportData.id.asc())
        .all()
    )

    merged: dict[str, dict] = {}  # month_str → 항목

    for row in all_raws:
        try:
            raw_data = json.loads(row.data_json)
            kpi_trend = raw_data.get("kpi_trend", {})
            trend_list = kpi_trend.get("trend", []) if isinstance(kpi_trend, dict) else kpi_trend
            if not trend_list:
                continue
            for t in trend_list:
                label = t.get("label", "")
                if not label:
                    continue
                month_str = _label_to_month(label)
                merged[month_str] = {
                    "month": month_str,
                    "followers_end": t.get("followers", 0),
                    "total_interactions": t.get("interactions", 0),
                    "total_reach": t.get("reach", 0),
                    "total_ad_spend": t.get("ad_spend", 0),
                }
        except Exception:
            continue

    if merged:
        return sorted(merged.values(), key=lambda x: x["month"])

    return []


@router.post("/sheet-list")
async def get_sheet_list(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """엑셀 파일의 시트 목록 반환"""
    file_bytes = await file.read()
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        sheets = []
        for i, name in enumerate(wb.sheetnames):
            sheets.append({"index": i, "name": name})
        wb.close()
        return {"sheets": sheets}
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"시트 목록 읽기 실패: {e}")


@router.post("/generate-excel-preview")
def generate_excel_preview(
    payload: dict = Body(...),
    current_user: User = Depends(get_current_user),
):
    buf = _build_excel_bytes(payload)
    month = payload.get("meta", {}).get("report_month", "data")
    client = payload.get("meta", {}).get("client", "NDG")
    filename = f"{client}_{month}_SNS데이터.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/{report_id}", response_model=ReportOut)
def get_report(report_id: int, db: Session = Depends(get_db),
               current_user: User = Depends(get_current_user)):
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다")
    return report


@router.delete("/{report_id}")
def delete_report(report_id: int, db: Session = Depends(get_db),
                  current_user: User = Depends(get_current_user)):
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다")
    db.delete(report)
    db.commit()
    return {"message": "삭제 완료"}


@router.post("/{report_id}/upload")
async def upload_excel(
    report_id: int,
    file: UploadFile = File(...),
    followers_start: int = Form(0),
    followers_end: int = Form(0),
    sheet_name: str = Form("0"),
    kpi_sheet_name: str = Form("none"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다")

    file_bytes = await file.read()

    # sheet_name 변환: 숫자면 int, 아니면 시트 이름(str)
    try:
        _sheet = int(sheet_name)
    except ValueError:
        _sheet = sheet_name

    # 컬럼 매핑 확인
    try:
        mapping_info = get_unmapped_columns(file_bytes, sheet_name=_sheet)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Excel 파일 읽기 실패: {e}")

    # 파싱
    try:
        parsed = parse_excel(
            file_bytes=file_bytes,
            report_month=report.report_month,
            client=report.client_name,
            followers_start=followers_start,
            followers_end=followers_end,
            sheet_name=_sheet,
        )
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    # KPI 추이 시트 파싱 (사용자 선택 시트, 없으면 스킵)
    if kpi_sheet_name == "none":
        kpi_trend = {}
    else:
        try:
            _kpi_sheet = int(kpi_sheet_name)
        except ValueError:
            _kpi_sheet = kpi_sheet_name
        kpi_trend = parse_kpi_sheet(file_bytes, sheet=_kpi_sheet)
    if kpi_trend:
        parsed["kpi_trend"] = kpi_trend

        # Sheet 2에서 팔로워 자동 보정 (수동 입력 없이)
        kpi_trend_list = kpi_trend.get("trend", [])
        if kpi_trend_list:
            try:
                m_label = str(int(report.report_month.split("-")[1])) + "월"
            except Exception:
                m_label = ""
            summary = parsed.get("current_month", {}).get("summary", {})
            for t in kpi_trend_list:
                if t.get("label") == m_label and t.get("followers", 0) > 0:
                    summary["followers_end"] = t["followers"]
                    break
            # 직전 월 팔로워를 followers_start로
            for t in reversed(kpi_trend_list):
                if t.get("label") != m_label and t.get("followers", 0) > 0:
                    summary["followers_start"] = t["followers"]
                    break
            # followers_end가 followers_start보다 작으면(잘못 계산된 경우) 재계산
            fs = summary.get("followers_start", 0)
            fe = summary.get("followers_end", 0)
            if fe < fs:
                summary["followers_end"] = fs + summary.get("total_new_followers", 0)

    # 저장
    data_record = ReportData(
        report_id=report_id,
        data_type="raw_excel",
        data_json=json.dumps(parsed, ensure_ascii=False),
    )
    db.add(data_record)
    report.status = "data_uploaded"
    report.updated_at = datetime.now(timezone.utc)
    db.commit()

    return {
        "message": "업로드 완료",
        "row_count": parsed["meta"]["row_count"],
        "mapping_info": mapping_info,
    }


@router.get("/{report_id}/data/{data_type}")
def get_report_data(report_id: int, data_type: str, db: Session = Depends(get_db),
                    current_user: User = Depends(get_current_user)):
    record = (
        db.query(ReportData)
        .filter(ReportData.report_id == report_id, ReportData.data_type == data_type)
        .order_by(ReportData.id.desc())
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail=f"{data_type} 데이터가 없습니다")
    return json.loads(record.data_json)


@router.put("/{report_id}/slides/{slide_number}")
def update_slide(
    report_id: int,
    slide_number: int,
    body: SlideUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # user_edited 우선, 없으면 agent_e_output
    record = None
    for dtype in ("user_edited", "agent_e_output"):
        record = (
            db.query(ReportData)
            .filter(ReportData.report_id == report_id, ReportData.data_type == dtype)
            .order_by(ReportData.id.desc())
            .first()
        )
        if record:
            break
    if not record:
        raise HTTPException(status_code=404, detail="보고서 구조 데이터가 없습니다")

    structure = json.loads(record.data_json)
    slides = structure.get("slides", [])
    target = next((s for s in slides if s["slide_number"] == slide_number), None)
    if not target:
        raise HTTPException(status_code=404, detail=f"슬라이드 {slide_number}을 찾을 수 없습니다")

    # 중첩 필드 업데이트 (dot notation + list index 지원)
    keys = body.field_key.split(".")
    obj = target["data"]
    original = None
    for k in keys[:-1]:
        if isinstance(obj, list):
            obj = obj[int(k)]
        else:
            obj = obj.setdefault(k, {})
    if isinstance(obj, list):
        original = obj[int(keys[-1])] if int(keys[-1]) < len(obj) else None
        obj[int(keys[-1])] = body.new_value
    else:
        original = obj.get(keys[-1])
        obj[keys[-1]] = body.new_value

    # 수정 이력 저장
    edit = SlideEdit(
        report_id=report_id,
        slide_number=slide_number,
        field_key=body.field_key,
        original_value=str(original) if original is not None else None,
        edited_value=body.new_value,
        edited_by=current_user.id,
    )
    db.add(edit)

    # user_edited 데이터로 저장
    new_record = ReportData(
        report_id=report_id,
        data_type="user_edited",
        data_json=json.dumps(structure, ensure_ascii=False),
    )
    db.add(new_record)
    db.commit()

    return {"message": "수정 저장 완료", "field_key": body.field_key}


@router.delete("/{report_id}/slides/{slide_number}")
def delete_slide(
    report_id: int,
    slide_number: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """슬라이드 삭제 후 user_edited로 저장"""
    record = None
    for dtype in ("user_edited", "agent_e_output"):
        record = (
            db.query(ReportData)
            .filter(ReportData.report_id == report_id, ReportData.data_type == dtype)
            .order_by(ReportData.id.desc())
            .first()
        )
        if record:
            break
    if not record:
        raise HTTPException(status_code=404, detail="보고서 구조 데이터가 없습니다")

    structure = json.loads(record.data_json)
    before = len(structure.get("slides", []))
    structure["slides"] = [s for s in structure.get("slides", []) if s["slide_number"] != slide_number]

    if len(structure["slides"]) == before:
        raise HTTPException(status_code=404, detail=f"슬라이드 {slide_number}을 찾을 수 없습니다")

    new_record = ReportData(
        report_id=report_id,
        data_type="user_edited",
        data_json=json.dumps(structure, ensure_ascii=False),
    )
    db.add(new_record)
    db.commit()
    return {"message": f"슬라이드 {slide_number} 삭제 완료"}


@router.get("/{report_id}/versions")
def list_versions(
    report_id: int,
    include_data: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    versions = (
        db.query(ReportVersion)
        .filter(ReportVersion.report_id == report_id)
        .order_by(ReportVersion.version_number.desc())
        .all()
    )
    result = []
    for v in versions:
        item = {
            "id": v.id,
            "version_number": v.version_number,
            "version_label": v.version_label,
            "created_at": v.created_at.isoformat(),
        }
        if include_data:
            item["data"] = json.loads(v.full_data_json)
        result.append(item)
    return result


@router.post("/{report_id}/versions")
def create_version(
    report_id: int,
    body: VersionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # 현재 최신 데이터 스냅샷
    record = (
        db.query(ReportData)
        .filter(
            ReportData.report_id == report_id,
            ReportData.data_type.in_(["user_edited", "agent_e_output"]),
        )
        .order_by(ReportData.id.desc())
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="저장할 데이터가 없습니다")

    latest_num = (
        db.query(ReportVersion)
        .filter(ReportVersion.report_id == report_id)
        .count()
    )
    version = ReportVersion(
        report_id=report_id,
        version_number=latest_num + 1,
        version_label=body.version_label,
        full_data_json=record.data_json,
        created_by=current_user.id,
    )
    db.add(version)
    db.commit()
    db.refresh(version)
    return {"message": "버전 저장 완료", "version_number": version.version_number}


@router.get("/{report_id}/versions/{version_number}")
def get_version(
    report_id: int,
    version_number: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    version = (
        db.query(ReportVersion)
        .filter(ReportVersion.report_id == report_id, ReportVersion.version_number == version_number)
        .first()
    )
    if not version:
        raise HTTPException(status_code=404, detail="버전을 찾을 수 없습니다")
    return {
        "version_number": version.version_number,
        "version_label": version.version_label,
        "created_at": version.created_at.isoformat(),
        "data": json.loads(version.full_data_json),
    }


@router.post("/{report_id}/versions/{version_number}/restore")
def restore_version(
    report_id: int,
    version_number: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """버전 스냅샷을 현재 상태로 복원 (user_edited로 저장)"""
    version = (
        db.query(ReportVersion)
        .filter(ReportVersion.report_id == report_id, ReportVersion.version_number == version_number)
        .first()
    )
    if not version:
        raise HTTPException(status_code=404, detail="버전을 찾을 수 없습니다")

    restored = ReportData(
        report_id=report_id,
        data_type="user_edited",
        data_json=version.full_data_json,
    )
    db.add(restored)
    db.commit()
    return {"message": f"v{version_number} 복원 완료"}


@router.put("/{report_id}/draft")
def save_draft(
    report_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """현재 최신 상태를 임시저장(draft 타입 upsert)"""
    record = (
        db.query(ReportData)
        .filter(
            ReportData.report_id == report_id,
            ReportData.data_type.in_(["user_edited", "agent_e_output"]),
        )
        .order_by(ReportData.id.desc())
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="저장할 데이터가 없습니다")

    draft = (
        db.query(ReportData)
        .filter(ReportData.report_id == report_id, ReportData.data_type == "draft")
        .first()
    )
    if draft:
        draft.data_json = record.data_json
    else:
        draft = ReportData(
            report_id=report_id,
            data_type="draft",
            data_json=record.data_json,
        )
        db.add(draft)
    db.commit()
    saved_at = datetime.now(timezone.utc).strftime("%H:%M")
    return {"message": "임시저장 완료", "saved_at": saved_at}


@router.post("/{report_id}/approve")
def approve_report(report_id: int, db: Session = Depends(get_db),
                   current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "editor"]:
        raise HTTPException(status_code=403, detail="승인 권한이 없습니다")
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다")
    report.status = "approved"
    report.updated_at = datetime.now(timezone.utc)
    db.commit()
    return {"message": "승인 완료"}


@router.post("/{report_id}/duplicate", response_model=ReportOut)
def duplicate_report(report_id: int, db: Session = Depends(get_db),
                     current_user: User = Depends(get_current_user)):
    original = db.query(Report).filter(Report.id == report_id).first()
    if not original:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다")
    new_report = Report(
        title=original.title + " (복사본)",
        client_name=original.client_name,
        report_month=original.report_month,
        status="draft",
        created_by=current_user.id,
    )
    db.add(new_report)
    db.commit()
    db.refresh(new_report)
    return new_report


@router.post("/{report_id}/direct-input")
def direct_input(
    report_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다")
    posts = payload.get("current_month", {}).get("posts", [])
    if "meta" not in payload:
        payload["meta"] = {}
    payload["meta"].setdefault("row_count", len(posts))
    payload["meta"].setdefault("client", report.client_name)
    payload["meta"].setdefault("report_month", report.report_month)
    payload["meta"]["parsed_at"] = datetime.now(timezone.utc).isoformat()
    data_record = ReportData(
        report_id=report_id,
        data_type="raw_excel",
        data_json=json.dumps(payload, ensure_ascii=False),
    )
    db.add(data_record)
    report.status = "data_uploaded"
    report.updated_at = datetime.now(timezone.utc)
    db.commit()
    return {"message": "데이터 저장 완료", "row_count": len(posts)}


@router.get("/{report_id}/export/excel")
def export_excel(report_id: int, db: Session = Depends(get_db),
                 current_user: User = Depends(get_current_user)):
    record = (
        db.query(ReportData)
        .filter(ReportData.report_id == report_id, ReportData.data_type == "raw_excel")
        .order_by(ReportData.id.desc())
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="데이터가 없습니다. 먼저 데이터를 입력하세요.")
    report = db.query(Report).filter(Report.id == report_id).first()
    data = json.loads(record.data_json)
    buf = _build_excel_bytes(data)
    filename = f"{report.client_name}_{report.report_month}_SNS데이터.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


# ── 슬라이드 이미지 업로드 ─────────────────────────────────────────

import uuid as _uuid
import shutil as _shutil

_IMG_DIR = Path(__file__).resolve().parent.parent.parent / "uploads" / "images"

@router.post("/{report_id}/slides/{slide_number}/image")
async def upload_slide_image(
    report_id: int,
    slide_number: int,
    file: UploadFile = File(...),
    field_key: str = Form("image_url"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="보고서를 찾을 수 없습니다")

    _IMG_DIR.mkdir(parents=True, exist_ok=True)
    ext = Path(file.filename).suffix or ".jpg"
    fname = f"{report_id}_{slide_number}_{_uuid.uuid4().hex[:8]}{ext}"
    fpath = _IMG_DIR / fname

    with open(fpath, "wb") as f:
        _shutil.copyfileobj(file.file, f)

    image_url = f"/static/uploads/images/{fname}"

    def _apply_image(data_json: str) -> str:
        data = json.loads(data_json)
        for slide in data.get("slides", []):
            if slide.get("slide_number") == slide_number:
                parts = field_key.split(".")
                target = slide["data"]
                for part in parts[:-1]:
                    if part.isdigit():
                        target = target[int(part)]
                    else:
                        target = target.setdefault(part, {})
                target[parts[-1]] = image_url
                break
        return json.dumps(data, ensure_ascii=False)

    # user_edited 우선, 없으면 agent_e_output에 저장
    for dtype in ("user_edited", "agent_e_output"):
        record = (
            db.query(ReportData)
            .filter(ReportData.report_id == report_id, ReportData.data_type == dtype)
            .order_by(ReportData.id.desc())
            .first()
        )
        if record:
            record.data_json = _apply_image(record.data_json)
            db.commit()
            break

    return {"image_url": image_url}
